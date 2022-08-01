import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time

# username = input("enter username: ")
# password = input("enter password: ")
# user = "pandyaharry777@gmail.com"
# pass = "Harry@786$"

# b = openpyxl.load_workbook("Book1.xlsx")
# sht = b.active

# username = sht.cell(row=2, column=1).value
# password = sht.cell(row=2, column=2).value

# driver = webdriver.Chrome(service=Service("chromedriver.exe"))
# driver.implicitly_wait(0.5)
# driver.maximize_window()

# driver.get('https://github.com/login')

# element1 = driver.find_element(By.NAME, 'login')
# element1.send_keys(username)

# element2 = driver.find_element(By.NAME, 'password')
# element2.send_keys(password)

# driver.find_element(By.NAME, 'commit').click()
# # for logout
# driver.find_element(
#     By.XPATH, "//*[@class='details-overlay details-reset js-feature-preview-indicator-container']").click()
# driver.find_element(
#     By.XPATH, "//button[contains(@class, 'dropdown-signout')]").click()

# time.sleep(2)
# driver.quit()



driver = webdriver.Chrome(service=Service("chromedriver.exe"))
driver.implicitly_wait(0.5)
driver.maximize_window()

driver.get('file:///C:/Users/Student/Desktop/index.html')

driver.find_element(By.XPATH, "//*[@type='text']").send_keys("hi")

time.sleep(5)
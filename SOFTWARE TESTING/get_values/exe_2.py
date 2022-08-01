import unittest
from openpyxl import load_workbook, cell


class Test(unittest.TestCase):
    def testReadExcelFile(self):
        wb = load_workbook("Book1.xlsx")
        sheet = wb.active
        # reading single value from Sheet
        b1 = sheet['B1']
        print("\nB1 --> ", b1.value)

        b2 = sheet['A1']
        print("A1 --> ", b2.value)

        b3 = sheet.cell(row=3, column=2)
        print("B3 --> ", b3.value)


if __name__ == '__main__':
    unittest.main(verbosity=2)

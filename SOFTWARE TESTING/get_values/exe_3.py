import unittest
from openpyxl import load_workbook, cell
import openpyxl


class Test(unittest.TestCase):
    def testReadExcelFile(self):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("Sheet2", 0)
        sheet["A1"].value = "This is Newly Created File"
        wb.save("Book2.xlsx")


if __name__ == '__main__':
    unittest.main(verbosity=2)

import unittest
from openpyxl import load_workbook, cell


class Test(unittest.TestCase):
    def testReadExcelFile(self):
        wb = load_workbook("Book1.xlsx")
        sheet = wb["Sheet1"]

        print("\nTotal no of rows", sheet.max_row)
        for cell in sheet['C']:
            print(cell.value)


if __name__ == '__main__':
    unittest.main(verbosity=2)
